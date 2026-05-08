"""Multi-format export logic for weather data.

This module provides exporters to convert weather records
into various formats (JSON, CSV, Excel) for download.
"""

import csv
import io
import json
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional
from xml.dom import minidom
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

# Try to import PDF library, but don't fail if not installed
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from app.models.weather import WeatherRecord
from app.repositories.weather_repository import WeatherRepository

logger = logging.getLogger(__name__)


class WeatherExporter:
    """Base exporter class with common utilities."""

    def __init__(self, db: AsyncSession):
        """Initialize exporter with database session.

        Args:
            db: SQLAlchemy async session
        """
        self.db = db
        self.repository = WeatherRepository(db)

    async def get_record_with_relations(
        self, record_id: int
    ) -> Optional[WeatherRecord]:
        """Fetch weather record with all related data.

        Args:
            record_id: ID of the weather record

        Returns:
            Weather record with relations if found, None otherwise
        """
        return await self.repository.get_weather_record_by_id(record_id)


class JSONExporter(WeatherExporter):
    """Export weather data to JSON format."""

    async def export(self, record_id: int) -> Optional[Dict[str, Any]]:
        """Export weather record to JSON-serializable dictionary.

        Args:
            record_id: ID of the weather record

        Returns:
            Dictionary representation of the record, or None if not found
        """
        record = await self.get_record_with_relations(record_id)
        if not record:
            logger.warning(f"Record {record_id} not found for JSON export")
            return None

        try:
            return self._serialize_record(record)
        except Exception as e:
            logger.error(f"Error exporting record {record_id} to JSON: {str(e)}")
            raise

    def _serialize_record(self, record: WeatherRecord) -> Dict[str, Any]:
        """Serialize a weather record to dictionary.

        Args:
            record: WeatherRecord instance

        Returns:
            Dictionary representation
        """
        # Base record data
        result = {
            "id": record.id,
            "location_name": record.location_name,
            "location_type": record.location_type,
            "latitude": float(record.latitude) if record.latitude else None,
            "longitude": float(record.longitude) if record.longitude else None,
            "start_date": record.start_date.isoformat() if record.start_date else None,
            "end_date": record.end_date.isoformat() if record.end_date else None,
            "user_notes": record.user_notes,
            "created_at": record.created_at.isoformat() if record.created_at else None,
            "updated_at": record.updated_at.isoformat() if record.updated_at else None,
        }

        # Weather data (daily forecasts)
        if record.weather_data:
            result["weather_data"] = [
                {
                    "forecast_date": wd.forecast_date.isoformat()
                    if wd.forecast_date
                    else None,
                    "temperature": wd.temperature,
                    "feels_like": wd.feels_like,
                    "temp_min": wd.temp_min,
                    "temp_max": wd.temp_max,
                    "humidity": wd.humidity,
                    "pressure": wd.pressure,
                    "wind_speed": wd.wind_speed,
                    "weather_description": wd.weather_description,
                    "icon_code": wd.icon_code,
                }
                for wd in record.weather_data
            ]

        # YouTube videos
        if record.youtube_videos:
            result["youtube_videos"] = [
                {
                    "video_id": yv.video_id,
                    "title": yv.title,
                    "description": yv.description,
                    "url": yv.url,
                    "thumbnail_url": yv.thumbnail_url,
                    "channel_title": yv.channel_title,
                    "published_at": yv.published_at,
                }
                for yv in record.youtube_videos
            ]

        # Map locations
        if record.map_locations:
            result["map_locations"] = [
                {
                    "place_id": ml.place_id,
                    "formatted_address": ml.formatted_address,
                    "map_url": ml.map_url,
                    "static_map_url": ml.static_map_url,
                    "latitude": float(ml.lat) if ml.lat else None,
                    "longitude": float(ml.lng) if ml.lng else None,
                    "place_type": ml.place_type,
                    "point_of_interest": ml.point_of_interest,
                }
                for ml in record.map_locations
            ]

        # Additional API data
        if record.additional_api_data:
            result["additional_api_data"] = [
                {
                    "api_name": ad.api_name,
                    "data_type": ad.data_type,
                    "payload": ad.payload,
                    "fetched_at": ad.fetched_at.isoformat() if ad.fetched_at else None,
                }
                for ad in record.additional_api_data
            ]

        return result


class CSVExporter(WeatherExporter):
    """Export weather data to CSV format."""

    async def export(self, record_id: int) -> Optional[str]:
        """Export weather record to CSV string.

        Args:
            record_id: ID of the weather record

        Returns:
            CSV string, or None if record not found
        """
        record = await self.get_record_with_relations(record_id)
        if not record:
            logger.warning(f"Record {record_id} not found for CSV export")
            return None

        try:
            output = io.StringIO()
            writer = csv.writer(output)

            # Write main record info
            writer.writerow(["Weather Record Export"])
            writer.writerow([])
            writer.writerow(["Record ID", record.id])
            writer.writerow(["Location Name", record.location_name])
            writer.writerow(["Location Type", record.location_type])
            writer.writerow(["Latitude", record.latitude])
            writer.writerow(["Longitude", record.longitude])
            writer.writerow(["Start Date", record.start_date])
            writer.writerow(["End Date", record.end_date])
            writer.writerow(["User Notes", record.user_notes or ""])
            writer.writerow([])

            # Write weather data
            if record.weather_data:
                writer.writerow(["Weather Forecast Data"])
                writer.writerow(
                    [
                        "Date",
                        "Temperature (°C)",
                        "Feels Like (°C)",
                        "Min Temp (°C)",
                        "Max Temp (°C)",
                        "Humidity (%)",
                        "Pressure (hPa)",
                        "Wind Speed (m/s)",
                        "Description",
                    ]
                )
                for wd in sorted(
                    record.weather_data, key=lambda x: x.forecast_date or ""
                ):
                    writer.writerow(
                        [
                            wd.forecast_date,
                            wd.temperature,
                            wd.feels_like,
                            wd.temp_min,
                            wd.temp_max,
                            wd.humidity,
                            wd.pressure,
                            wd.wind_speed,
                            wd.weather_description or "",
                        ]
                    )
                writer.writerow([])

            # Write YouTube videos
            if record.youtube_videos:
                writer.writerow(["YouTube Videos"])
                writer.writerow(["Title", "Channel", "URL", "Published At"])
                for yv in record.youtube_videos:
                    writer.writerow(
                        [
                            yv.title or "",
                            yv.channel_title or "",
                            yv.url or "",
                            yv.published_at or "",
                        ]
                    )
                writer.writerow([])

            # Write map locations
            if record.map_locations:
                writer.writerow(["Map Locations"])
                writer.writerow(
                    ["Address", "Map URL", "Place Type", "Point of Interest"]
                )
                for ml in record.map_locations:
                    writer.writerow(
                        [
                            ml.formatted_address or "",
                            ml.map_url or "",
                            ml.place_type or "",
                            ml.point_of_interest or "",
                        ]
                    )

            return output.getvalue()

        except Exception as e:
            logger.error(f"Error exporting record {record_id} to CSV: {str(e)}")
            raise


class XMLExporter(WeatherExporter):
    """Export weather data to XML format."""

    async def export(self, record_id: int) -> Optional[str]:
        """Export weather record to XML string.

        Args:
            record_id: ID of the weather record

        Returns:
            XML string, or None if record not found
        """
        record = await self.get_record_with_relations(record_id)
        if not record:
            logger.warning(f"Record {record_id} not found for XML export")
            return None

        try:
            # Create root element
            root = ET.Element("WeatherRecordExport")

            # Add record info
            record_elem = ET.SubElement(root, "Record")
            ET.SubElement(record_elem, "Id").text = str(record.id)
            ET.SubElement(record_elem, "LocationName").text = record.location_name
            ET.SubElement(record_elem, "LocationType").text = record.location_type
            ET.SubElement(record_elem, "Latitude").text = (
                str(record.latitude) if record.latitude else ""
            )
            ET.SubElement(record_elem, "Longitude").text = (
                str(record.longitude) if record.longitude else ""
            )
            ET.SubElement(record_elem, "StartDate").text = (
                record.start_date.isoformat() if record.start_date else ""
            )
            ET.SubElement(record_elem, "EndDate").text = (
                record.end_date.isoformat() if record.end_date else ""
            )
            ET.SubElement(record_elem, "UserNotes").text = record.user_notes or ""
            ET.SubElement(record_elem, "CreatedAt").text = (
                record.created_at.isoformat() if record.created_at else ""
            )
            ET.SubElement(record_elem, "UpdatedAt").text = (
                record.updated_at.isoformat() if record.updated_at else ""
            )

            # Add weather data
            if record.weather_data:
                weather_elem = ET.SubElement(root, "WeatherData")
                for wd in sorted(
                    record.weather_data, key=lambda x: x.forecast_date or ""
                ):
                    day_elem = ET.SubElement(weather_elem, "Day")
                    ET.SubElement(day_elem, "ForecastDate").text = (
                        wd.forecast_date.isoformat() if wd.forecast_date else ""
                    )
                    ET.SubElement(day_elem, "Temperature").text = (
                        str(wd.temperature) if wd.temperature else ""
                    )
                    ET.SubElement(day_elem, "FeelsLike").text = (
                        str(wd.feels_like) if wd.feels_like else ""
                    )
                    ET.SubElement(day_elem, "TempMin").text = (
                        str(wd.temp_min) if wd.temp_min else ""
                    )
                    ET.SubElement(day_elem, "TempMax").text = (
                        str(wd.temp_max) if wd.temp_max else ""
                    )
                    ET.SubElement(day_elem, "Humidity").text = (
                        str(wd.humidity) if wd.humidity else ""
                    )
                    ET.SubElement(day_elem, "Pressure").text = (
                        str(wd.pressure) if wd.pressure else ""
                    )
                    ET.SubElement(day_elem, "WindSpeed").text = (
                        str(wd.wind_speed) if wd.wind_speed else ""
                    )
                    ET.SubElement(day_elem, "Description").text = (
                        wd.weather_description or ""
                    )
                    ET.SubElement(day_elem, "IconCode").text = wd.icon_code or ""

            # Add YouTube videos
            if record.youtube_videos:
                videos_elem = ET.SubElement(root, "YouTubeVideos")
                for yv in record.youtube_videos:
                    video_elem = ET.SubElement(videos_elem, "Video")
                    ET.SubElement(video_elem, "VideoId").text = yv.video_id or ""
                    ET.SubElement(video_elem, "Title").text = yv.title or ""
                    ET.SubElement(video_elem, "Description").text = yv.description or ""
                    ET.SubElement(video_elem, "URL").text = yv.url or ""
                    ET.SubElement(video_elem, "ThumbnailURL").text = (
                        yv.thumbnail_url or ""
                    )
                    ET.SubElement(video_elem, "ChannelTitle").text = (
                        yv.channel_title or ""
                    )
                    ET.SubElement(video_elem, "PublishedAt").text = (
                        yv.published_at or ""
                    )

            # Add map locations
            if record.map_locations:
                maps_elem = ET.SubElement(root, "MapLocations")
                for ml in record.map_locations:
                    loc_elem = ET.SubElement(maps_elem, "Location")
                    ET.SubElement(loc_elem, "PlaceId").text = ml.place_id or ""
                    ET.SubElement(loc_elem, "FormattedAddress").text = (
                        ml.formatted_address or ""
                    )
                    ET.SubElement(loc_elem, "MapURL").text = ml.map_url or ""
                    ET.SubElement(loc_elem, "StaticMapURL").text = (
                        ml.static_map_url or ""
                    )
                    ET.SubElement(loc_elem, "Latitude").text = (
                        str(ml.lat) if ml.lat else ""
                    )
                    ET.SubElement(loc_elem, "Longitude").text = (
                        str(ml.lng) if ml.lng else ""
                    )
                    ET.SubElement(loc_elem, "PlaceType").text = ml.place_type or ""
                    ET.SubElement(loc_elem, "PointOfInterest").text = (
                        ml.point_of_interest or ""
                    )

            # Add additional API data
            if record.additional_api_data:
                api_elem = ET.SubElement(root, "AdditionalAPIData")
                for ad in record.additional_api_data:
                    data_elem = ET.SubElement(api_elem, "Data")
                    ET.SubElement(data_elem, "APIName").text = ad.api_name or ""
                    ET.SubElement(data_elem, "DataType").text = ad.data_type or ""
                    ET.SubElement(data_elem, "Payload").text = (
                        json.dumps(ad.payload) if ad.payload else ""
                    )
                    ET.SubElement(data_elem, "FetchedAt").text = (
                        ad.fetched_at.isoformat() if ad.fetched_at else ""
                    )

            # Pretty print XML
            xml_str = ET.tostring(root, encoding="utf-8")
            dom = minidom.parseString(xml_str)
            return dom.toprettyxml(indent="  ")

        except Exception as e:
            logger.error(f"Error exporting record {record_id} to XML: {str(e)}")
            raise


class PDFExporter(WeatherExporter):
    """Export weather data to PDF format."""

    async def export(self, record_id: int) -> Optional[bytes]:
        """Export weather record to PDF bytes.

        Args:
            record_id: ID of the weather record

        Returns:
            PDF file as bytes, or None if record not found
        """
        if not REPORTLAB_AVAILABLE:
            logger.error("reportlab not installed. Cannot export to PDF.")
            raise ImportError(
                "PDF export requires reportlab. Install it with: pip install reportlab"
            )

        record = await self.get_record_with_relations(record_id)
        if not record:
            logger.warning(f"Record {record_id} not found for PDF export")
            return None

        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title = Paragraph(
                f"Weather Record: {record.location_name}", styles["Title"]
            )
            story.append(title)
            story.append(Spacer(1, 12))

            # Record info
            info_data = [
                ["Record ID:", str(record.id)],
                ["Location:", record.location_name],
                ["Type:", record.location_type],
                [
                    "Start Date:",
                    record.start_date.isoformat() if record.start_date else "N/A",
                ],
                [
                    "End Date:",
                    record.end_date.isoformat() if record.end_date else "N/A",
                ],
                ["Notes:", record.user_notes or "N/A"],
            ]
            info_table = Table(info_data, colWidths=[100, 300])
            info_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.lightgrey),
                        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            story.append(info_table)
            story.append(Spacer(1, 12))

            # Weather data table
            if record.weather_data:
                story.append(Paragraph("Weather Forecast Data", styles["Heading2"]))
                weather_data = [
                    [
                        "Date",
                        "Temp (°C)",
                        "Feels Like",
                        "Min",
                        "Max",
                        "Humidity",
                        "Description",
                    ]
                ]
                for wd in sorted(
                    record.weather_data, key=lambda x: x.forecast_date or ""
                ):
                    weather_data.append(
                        [
                            wd.forecast_date.isoformat() if wd.forecast_date else "",
                            str(wd.temperature) if wd.temperature else "",
                            str(wd.feels_like) if wd.feels_like else "",
                            str(wd.temp_min) if wd.temp_min else "",
                            str(wd.temp_max) if wd.temp_max else "",
                            str(wd.humidity) if wd.humidity else "",
                            wd.weather_description or "",
                        ]
                    )
                weather_table = Table(weather_data)
                weather_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )
                story.append(weather_table)
                story.append(Spacer(1, 12))

            # YouTube videos
            if record.youtube_videos:
                story.append(Paragraph("YouTube Videos", styles["Heading2"]))
                for yv in record.youtube_videos:
                    video_text = f"<b>{yv.title or 'Untitled'}</b><br/>"
                    video_text += f"Channel: {yv.channel_title or 'Unknown'}<br/>"
                    video_text += f"URL: {yv.url or 'N/A'}"
                    story.append(Paragraph(video_text, styles["Normal"]))
                    story.append(Spacer(1, 6))

            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Error exporting record {record_id} to PDF: {str(e)}")
            raise


class MarkdownExporter(WeatherExporter):
    """Export weather data to Markdown format."""

    async def export(self, record_id: int) -> Optional[str]:
        """Export weather record to Markdown string.

        Args:
            record_id: ID of the weather record

        Returns:
            Markdown string, or None if record not found
        """
        record = await self.get_record_with_relations(record_id)
        if not record:
            logger.warning(f"Record {record_id} not found for Markdown export")
            return None

        try:
            md_lines = []

            # Title
            md_lines.append(f"# Weather Record: {record.location_name}")
            md_lines.append("")

            # Record info
            md_lines.append("## Record Information")
            md_lines.append("")
            md_lines.append(f"- **ID:** {record.id}")
            md_lines.append(f"- **Location:** {record.location_name}")
            md_lines.append(f"- **Type:** {record.location_type}")
            md_lines.append(
                f"- **Coordinates:** ({record.latitude}, {record.longitude})"
            )
            md_lines.append(
                f"- **Start Date:** {record.start_date.isoformat() if record.start_date else 'N/A'}"
            )
            md_lines.append(
                f"- **End Date:** {record.end_date.isoformat() if record.end_date else 'N/A'}"
            )
            md_lines.append(f"- **Notes:** {record.user_notes or 'N/A'}")
            md_lines.append(
                f"- **Created:** {record.created_at.isoformat() if record.created_at else 'N/A'}"
            )
            md_lines.append(
                f"- **Updated:** {record.updated_at.isoformat() if record.updated_at else 'N/A'}"
            )
            md_lines.append("")

            # Weather data
            if record.weather_data:
                md_lines.append("## Weather Forecast Data")
                md_lines.append("")
                md_lines.append(
                    "| Date | Temperature (°C) | Feels Like (°C) | Min (°C) | Max (°C) | Humidity (%) | Description |"
                )
                md_lines.append(
                    "|------|------------------|-----------------|-----------|-----------|---------------|-------------|"
                )
                for wd in sorted(
                    record.weather_data, key=lambda x: x.forecast_date or ""
                ):
                    md_lines.append(
                        f"| {wd.forecast_date.isoformat() if wd.forecast_date else 'N/A'} "
                        f"| {wd.temperature if wd.temperature else 'N/A'} "
                        f"| {wd.feels_like if wd.feels_like else 'N/A'} "
                        f"| {wd.temp_min if wd.temp_min else 'N/A'} "
                        f"| {wd.temp_max if wd.temp_max else 'N/A'} "
                        f"| {wd.humidity if wd.humidity else 'N/A'} "
                        f"| {wd.weather_description or 'N/A'} |"
                    )
                md_lines.append("")

            # YouTube videos
            if record.youtube_videos:
                md_lines.append("## YouTube Videos")
                md_lines.append("")
                for yv in record.youtube_videos:
                    md_lines.append(f"### {yv.title or 'Untitled'}")
                    md_lines.append(f"- **Channel:** {yv.channel_title or 'Unknown'}")
                    md_lines.append(f"- **URL:** {yv.url or 'N/A'}")
                    md_lines.append(f"- **Published:** {yv.published_at or 'N/A'}")
                    if yv.description:
                        md_lines.append(f"- **Description:** {yv.description}")
                    md_lines.append("")

            # Map locations
            if record.map_locations:
                md_lines.append("## Map Locations")
                md_lines.append("")
                for ml in record.map_locations:
                    md_lines.append(f"### {ml.formatted_address or 'Unknown Location'}")
                    md_lines.append(f"- **Place ID:** {ml.place_id or 'N/A'}")
                    md_lines.append(f"- **Type:** {ml.place_type or 'N/A'}")
                    md_lines.append(f"- **Map URL:** {ml.map_url or 'N/A'}")
                    md_lines.append("")

            # Additional API data
            if record.additional_api_data:
                md_lines.append("## Additional API Data")
                md_lines.append("")
                for ad in record.additional_api_data:
                    md_lines.append(f"### {ad.api_name or 'Unknown API'}")
                    md_lines.append(f"- **Data Type:** {ad.data_type or 'N/A'}")
                    md_lines.append(
                        f"- **Fetched At:** {ad.fetched_at.isoformat() if ad.fetched_at else 'N/A'}"
                    )
                    if ad.payload:
                        md_lines.append("**Payload:**")
                        md_lines.append("```json")
                        md_lines.append(json.dumps(ad.payload, indent=2))
                        md_lines.append("```")
                    md_lines.append("")

            return "\n".join(md_lines)

        except Exception as e:
            logger.error(f"Error exporting record {record_id} to Markdown: {str(e)}")
            raise


class ExcelExporter(WeatherExporter):
    """Export weather data to Excel format."""

    async def export(self, record_id: int) -> Optional[bytes]:
        """Export weather record to Excel file bytes.

        Args:
            record_id: ID of the weather record

        Returns:
            Excel file as bytes, or None if record not found
        """
        record = await self.get_record_with_relations(record_id)
        if not record:
            logger.warning(f"Record {record_id} not found for Excel export")
            return None

        try:
            workbook = openpyxl.Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)

            # Create sheets
            self._create_summary_sheet(workbook, record)
            if record.weather_data:
                self._create_weather_data_sheet(workbook, record)
            if record.youtube_videos:
                self._create_youtube_sheet(workbook, record)
            if record.map_locations:
                self._create_map_sheet(workbook, record)

            # Save to bytes
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            return excel_buffer.read()

        except Exception as e:
            logger.error(f"Error exporting record {record_id} to Excel: {str(e)}")
            raise

    def _create_summary_sheet(self, workbook: openpyxl.Workbook, record: WeatherRecord):
        """Create summary sheet with record details.

        Args:
            workbook: openpyxl Workbook instance
            record: WeatherRecord instance
        """
        sheet = workbook.create_sheet("Summary")

        # Header style
        header_font = Font(bold=True, size=14)
        label_font = Font(bold=True)

        # Title
        sheet["A1"] = "Weather Record Export"
        sheet["A1"].font = header_font
        sheet["A1"].fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )

        # Record details
        details = [
            ("Record ID", record.id),
            ("Location Name", record.location_name),
            ("Location Type", record.location_type),
            ("Latitude", record.latitude),
            ("Longitude", record.longitude),
            ("Start Date", record.start_date),
            ("End Date", record.end_date),
            ("User Notes", record.user_notes or ""),
            ("Created At", record.created_at),
            ("Updated At", record.updated_at),
        ]

        for idx, (label, value) in enumerate(details, start=3):
            cell_label = sheet[f"A{idx}"]
            cell_label.value = label
            cell_label.font = label_font

            cell_value = sheet[f"B{idx}"]
            cell_value.value = value

    def _create_weather_data_sheet(
        self, workbook: openpyxl.Workbook, record: WeatherRecord
    ):
        """Create weather data sheet.

        Args:
            workbook: openpyxl Workbook instance
            record: WeatherRecord instance
        """
        sheet = workbook.create_sheet("Weather Data")

        # Headers
        headers = [
            "Date",
            "Temperature (°C)",
            "Feels Like (°C)",
            "Min Temp (°C)",
            "Max Temp (°C)",
            "Humidity (%)",
            "Pressure (hPa)",
            "Wind Speed (m/s)",
            "Description",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
            )

        # Data rows
        for row_idx, wd in enumerate(
            sorted(record.weather_data, key=lambda x: x.forecast_date or ""), start=2
        ):
            sheet.cell(row=row_idx, column=1).value = wd.forecast_date
            sheet.cell(row=row_idx, column=2).value = wd.temperature
            sheet.cell(row=row_idx, column=3).value = wd.feels_like
            sheet.cell(row=row_idx, column=4).value = wd.temp_min
            sheet.cell(row=row_idx, column=5).value = wd.temp_max
            sheet.cell(row=row_idx, column=6).value = wd.humidity
            sheet.cell(row=row_idx, column=7).value = wd.pressure
            sheet.cell(row=row_idx, column=8).value = wd.wind_speed
            sheet.cell(row=row_idx, column=9).value = wd.weather_description

    def _create_youtube_sheet(self, workbook: openpyxl.Workbook, record: WeatherRecord):
        """Create YouTube videos sheet.

        Args:
            workbook: openpyxl Workbook instance
            record: WeatherRecord instance
        """
        sheet = workbook.create_sheet("YouTube Videos")

        # Headers
        headers = ["Title", "Channel", "URL", "Published At"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
            )

        # Data rows
        for row_idx, yv in enumerate(record.youtube_videos, start=2):
            sheet.cell(row=row_idx, column=1).value = yv.title
            sheet.cell(row=row_idx, column=2).value = yv.channel_title
            sheet.cell(row=row_idx, column=3).value = yv.url
            sheet.cell(row=row_idx, column=4).value = yv.published_at

    def _create_map_sheet(self, workbook: openpyxl.Workbook, record: WeatherRecord):
        """Create map locations sheet.

        Args:
            workbook: openpyxl Workbook instance
            record: WeatherRecord instance
        """
        sheet = workbook.create_sheet("Map Locations")

        # Headers
        headers = ["Address", "Map URL", "Place Type", "Point of Interest"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
            )

        # Data rows
        for row_idx, ml in enumerate(record.map_locations, start=2):
            sheet.cell(row=row_idx, column=1).value = ml.formatted_address
            sheet.cell(row=row_idx, column=2).value = ml.map_url
            sheet.cell(row=row_idx, column=3).value = ml.place_type
            sheet.cell(row=row_idx, column=4).value = ml.point_of_interest


class ExportManager:
    """Manager to handle all export formats with a unified interface."""

    def __init__(self, db: AsyncSession):
        """Initialize export manager with database session.

        Args:
            db: SQLAlchemy async session
        """
        self.db = db

    async def export_to_json(self, record_id: int) -> Optional[Dict[str, Any]]:
        """Export record to JSON format.

        Args:
            record_id: ID of the weather record

        Returns:
            JSON-serializable dictionary
        """
        try:
            exporter = JSONExporter(self.db)
            return await exporter.export(record_id)
        except Exception as e:
            logger.error(f"Error in ExportManager.export_to_json: {str(e)}")
            raise

    async def export_to_csv(self, record_id: int) -> Optional[str]:
        """Export record to CSV format.

        Args:
            record_id: ID of the weather record

        Returns:
            CSV string
        """
        try:
            exporter = CSVExporter(self.db)
            return await exporter.export(record_id)
        except Exception as e:
            logger.error(f"Error in ExportManager.export_to_csv: {str(e)}")
            raise

    async def export_to_xml(self, record_id: int) -> Optional[str]:
        """Export record to XML format.

        Args:
            record_id: ID of the weather record

        Returns:
            XML string
        """
        try:
            exporter = XMLExporter(self.db)
            return await exporter.export(record_id)
        except Exception as e:
            logger.error(f"Error in ExportManager.export_to_xml: {str(e)}")
            raise

    async def export_to_pdf(self, record_id: int) -> Optional[bytes]:
        """Export record to PDF format.

        Args:
            record_id: ID of the weather record

        Returns:
            PDF file as bytes
        """
        try:
            exporter = PDFExporter(self.db)
            return await exporter.export(record_id)
        except Exception as e:
            logger.error(f"Error in ExportManager.export_to_pdf: {str(e)}")
            raise

    async def export_to_markdown(self, record_id: int) -> Optional[str]:
        """Export record to Markdown format.

        Args:
            record_id: ID of the weather record

        Returns:
            Markdown string
        """
        try:
            exporter = MarkdownExporter(self.db)
            return await exporter.export(record_id)
        except Exception as e:
            logger.error(f"Error in ExportManager.export_to_markdown: {str(e)}")
            raise

    async def export_to_excel(self, record_id: int) -> Optional[bytes]:
        """Export record to Excel format.

        Args:
            record_id: ID of the weather record

        Returns:
            Excel file as bytes
        """
        try:
            exporter = ExcelExporter(self.db)
            return await exporter.export(record_id)
        except Exception as e:
            logger.error(f"Error in ExportManager.export_to_excel: {str(e)}")
            raise

    async def get_export_filename(self, record_id: int, format: str) -> Optional[str]:
        """Generate export filename for a record.

        Args:
            record_id: ID of the weather record
            format: Export format (json, csv, xml, pdf, md, xlsx)

        Returns:
            Filename string, or None if record not found
        """
        exporter = WeatherExporter(self.db)
        record = await exporter.get_record_with_relations(record_id)

        if not record:
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_location = "".join(
            c for c in record.location_name if c.isalnum() or c in (" ", "-", "_")
        ).strip()
        safe_location = safe_location.replace(" ", "_")[:50]  # Limit length

        extension_map = {
            "json": "json",
            "csv": "csv",
            "xml": "xml",
            "pdf": "pdf",
            "md": "md",
            "xlsx": "xlsx",
        }

        extension = extension_map.get(format.lower(), "txt")
        return f"weather_{safe_location}_{record_id}_{timestamp}.{extension}"

    async def export(self, record_id: int, format: str) -> Optional[Any]:
        """Unified export method that routes to appropriate exporter.

        Args:
            record_id: ID of the weather record
            format: Export format (json, csv, xml, pdf, md, xlsx)

        Returns:
            Exported data in the appropriate format
        """
        format_lower = format.lower()

        if format_lower == "json":
            return await self.export_to_json(record_id)
        elif format_lower == "csv":
            return await self.export_to_csv(record_id)
        elif format_lower == "xml":
            return await self.export_to_xml(record_id)
        elif format_lower == "pdf":
            return await self.export_to_pdf(record_id)
        elif format_lower == "md" or format_lower == "markdown":
            return await self.export_to_markdown(record_id)
        elif format_lower == "xlsx" or format_lower == "excel":
            return await self.export_to_excel(record_id)
        else:
            raise ValueError(f"Unsupported export format: {format}")
